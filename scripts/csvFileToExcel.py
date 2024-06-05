import sys
import csv
import os.path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker,OneCellAnchor
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU,BASE_COL_WIDTH,pixels_to_points

if len(sys.argv) < 2:
    print("Usage: python script.py <input_file.csv>")
    sys.exit(1)

input_file = sys.argv[1]

# 读取 CSV 文件
with open(input_file, 'r', encoding='utf-8') as file:
    reader = csv.reader(file, delimiter='|')
    data = list(reader)

# 创建 Excel 工作簿和工作表
workbook = Workbook()
worksheet = workbook.active

# 设置图片列的列号
image_col_num = 2

# 设置图片大小
image_height = 100
image_width = 100
# 设置单元格行高
row_height_for_image = 100
row_height_for_text = 20

image_size = XDRPositiveSize2D(
    cx=pixels_to_EMU(image_width ),  # 设置图片宽度
    cy=pixels_to_EMU(image_height)  # 设置图片高度
)

# 计算包含中文、数字和英文的单元格宽度
def calculate_cell_width_by_str(char_list: str = ""):
    width = 0
    for char in char_list:
        if '\u4e00' <= char <= '\u9fff':  # 判断是否为中文字符
            width += 2
        elif char.isdigit() or char.isalpha():  # 判断是否为数字或英文字符
            width += 1
        else:  # 其他字符
            width += 1
    return width # * 1.2 + 2 # 默认字符宽度乘数 中英文计算会不正常

#遍历 data 列表中的每一行。行号从 1 开始,而不是默认的 0。
cell_width_list = []
for row_num, row_list in enumerate(data, start=1):
    cell_height = row_height_for_text

    #用于遍历当前行中的每一个单元格。列号从 1 开始,而不是默认的 0
    for col_num, cell_value in enumerate(row_list, start=1):
        cell_width = 2 # 默认两个字符宽度
        if col_num == image_col_num:  # 第四列是图片文件地址
            cell_image_paths = cell_value.split(',')
            offset_x = 0 #
            offset_y = 0
            cell_value_path = ""
            for image_path in cell_image_paths:
                if os.path.exists(image_path.strip()):
                    # 创建图片对象并设置大小
                    image = Image(image_path.strip())
                    image.height = image_height
                    image.width = image_width
                    marker = AnchorMarker(col=image_col_num-1,colOff=pixels_to_EMU(offset_y),row = row_num-1,rowOff=offset_y)
                    image.anchor = OneCellAnchor(_from = marker,ext = image_size )
                    # 将图片添加到工作表中，长度+1
                    worksheet.add_image(image)

                    cell_height = image_height
                    offset_y = offset_y + image_width
                else:
                    # 如果图片地址不存在,则显示文件名,长度+1
                    cell_value_path = cell_value_path + image_path
                    worksheet.cell(row=row_num, column=col_num, value=cell_value_path)
            # 自适应调整单元格大小,取最宽的列
            cell_width = offset_y / BASE_COL_WIDTH
        else:
            # 写入其他列的数据
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
            # 自适应调整单元格大小
            cell_width = calculate_cell_width_by_str(cell_value) 
        if(len(cell_width_list) < col_num ):
            cell_width_list.append(cell_width)
        elif(cell_width > cell_width_list[col_num-1]):
            cell_width_list[col_num-1] = cell_width

    worksheet.row_dimensions[row_num].height = pixels_to_points(cell_height) # * 0.75  # 转换为 Excel 的行高
# 设置所有列的列宽
for col_num, width in enumerate(cell_width_list, start=1):
    worksheet.column_dimensions[get_column_letter(col_num)].width =  width

# 保存 Excel 文件
output_file = input_file.replace('.csv', '.xlsx')
workbook.save(output_file)
print(f"CSV file '{input_file}' exported to '{output_file}'.")
